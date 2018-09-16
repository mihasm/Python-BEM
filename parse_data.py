__author__ = "Miha Smrekar"
__credits__ = ["Miha Smrekar"]
__license__ = "GPL"
__version__ = "0.2.3"
__maintainer__ = "Miha Smrekar"
__email__ = "miha.smrekar9@gmail.com"
__status__ = "Development"

from openpyxl import load_workbook
import numpy
from scipy import interpolate
from scipy.ndimage.interpolation import shift

try:
    wb = load_workbook("program_v4.xlsx")
    ws = wb["geometrija"]
    ws2 = wb["podatki"]
except FileNotFoundError:
    print(
        "File 'program_v4.xlsx' was not found, please rename your excel file to 'program_v3.xlsm' and move it to the "
        "folder of this file."
    )


def parse_podatki():
    """
    Parses turbine geometry data from excel file program_v3.xlsm.
    Extracts cL and cD curve data.
    :return: Returns dictionary with resulting points as arrays.
    """

    COLUMN_CL_X = 1
    COLUMN_CL_Y = 2
    COLUMN_CD_X = 4
    COLUMN_CD_Y = 5

    start_row = 2

    cL_x = []
    cL_y = []
    cD_x = []
    cD_y = []

    i = start_row

    while True:
        value_x = ws2.cell(row=i, column=COLUMN_CL_X).value
        value_y = ws2.cell(row=i, column=COLUMN_CL_Y).value

        if value_x == None and value_y == None:
            break
        else:
            cL_x.append(float(value_x))
            cL_y.append(float(value_y))
            i += 1

    i = start_row

    while True:
        value_x = ws2.cell(row=i, column=COLUMN_CD_X).value
        value_y = ws2.cell(row=i, column=COLUMN_CD_Y).value

        if value_x == None and value_y == None:
            break
        else:
            cD_x.append(float(value_x))
            cD_y.append(float(value_y))
            i += 1

    return {
        "cL_x": numpy.array(cL_x),
        "cL_y": numpy.array(cL_y),
        "cD_x": numpy.array(cD_x),
        "cD_y": numpy.array(cD_y),
    }


def parse_sections():
    """
    Parses turbine geometry data from excel file program_v3.xlsm.
    :return: Returns section radiuses, chord lengths, chord angles and section heights.
    """
    r = []
    c = []
    theta = []
    dr = []

    i = 4
    while True:
        _r = ws.cell(row=i, column=1).value
        _c = ws.cell(row=i, column=2).value
        _theta = ws.cell(row=i, column=3).value
        _dr = ws.cell(row=i, column=4).value
        if _r != None and _c != None and _theta != None and _dr != None:
            r.append(_r)
            c.append(_c)
            theta.append(_theta)
            dr.append(_dr)
        elif _r != None or _c != None or _theta != None or _dr != None:
            pass
        else:
            break

        i += 1

    r = numpy.array(r) * 1e-3  #
    c = numpy.array(c) * 1e-3  # m
    theta = numpy.array(theta)
    dr = numpy.array(dr) * 1e-3  # m

    return r, c, theta, dr
