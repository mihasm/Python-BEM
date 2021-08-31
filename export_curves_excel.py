import openpyxl
import json
import numpy as np

FILE = r"C:\Users\Miha\Google Drive\faks\BEM program\Projects\APC Thin Electric 10x7 propeller\APC_10x7E.bem"
OUT = "test.xlsx"

with open(FILE,"r") as f:
	file = json.loads(f.read())

i = 1
j = 1
#print()

wb = openpyxl.Workbook()
ws = wb.active

for foil,data in file["airfoils"].items():
	curves = data["gathered_curves"]
	i+=1
	e=ws.cell(row=i,column=1)
	e.value = "Airfoil=%s" % foil
	i+=1
	e=ws.cell(row=i,column=1)
	e.value = "Re"
	e=ws.cell(row=i,column=2)
	e.value = "Ncrit"
	e=ws.cell(row=i,column=3)
	e.value = "alpha"
	e=ws.cell(row=i,column=4)
	e.value = "Cl"
	e=ws.cell(row=i,column=5)
	e.value = "Cd"
	for r in curves:
		i+=1
		for j in range(1,6):
			e=ws.cell(row=i,column=j)
			e.value=r[j-1]
		print(r)

print(i)

wb.save(OUT)